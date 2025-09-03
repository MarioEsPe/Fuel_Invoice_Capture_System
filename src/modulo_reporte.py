"""
Module: modulo_reporte.py
Handles final report generation by updating report fields in Excel
and exporting the worksheet to PDF.

Functions:
    - generar_reporte(filename, sheet, turno, responsable, fecha_descarga, total_cantidad_natural, total_cantidad_20c)
"""

from openpyxl import load_workbook
import win32com.client as win32
import os
import pythoncom

def generar_reporte(
    filename, 
    sheet,
    turno, 
    responsable, 
    fecha_descarga, 
    total_cantidad_natural, 
    total_cantidad_20c, 
    output_dir="C:\\Users\\mario\\Desktop"
    ):
    """
    Finalize the report in Excel and export it to PDF.

    Args:
        - filename (str): Path to the Excel file.
        - sheet (str): Target worksheet name.
        - turno (str): Work Shift.
        - responsable (str): Responsible person for the report.
        - fecha_descarga (str): Date of operation.
        - total_cantidad_natural (float): Total liters at natural conditions.
        - total_cantidad_20c (float): Total liters at 20°C.
        - output_dir (str): Directory to save the generated PDF.

    Returns:
        - str: Path to the exported PDF file.
    """
    
    wb = load_workbook(filename)
    ws = wb[sheet]

    ws["D9"] = fecha_descarga
    ws["C11"] = turno
    ws["J32"] = responsable
    ws["F26"] = total_cantidad_natural
    ws["G26"] = total_cantidad_20c
    
    wb.save(filename)
    print("✅ Report finalized in Excel")
    
    # -----   Export to PDF   -----
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    
    abs_path = os.path.abspath(filename)
    wb_excel = excel.Workbooks.Open(abs_path)
    ws_excel = wb_excel.Worksheets[sheet]
    
    # Modify sheet orientation
    ws_excel.PageSetup.Orientation = 2  # 1 = Portrait, 2 = Landscape

    # Dynamic PDF name
    turno_clean = turno.replace("Turno", "").strip()
    pdf_name = f"Reporte_{fecha_descarga.replace('/', '-')}_Turno{turno_clean.replace(':', '')}.pdf"
    pdf_path = os.path.join(output_dir, pdf_name)
    
    ws_excel.ExportAsFixedFormat(0, pdf_path)
    
    wb_excel.Close(SaveChanges=False)
    excel.Quit()
    del wb_excel
    del excel
    pythoncom.CoUninitialize()

    print(f"✅ Report exported to PDF: {pdf_path}")

    return pdf_path