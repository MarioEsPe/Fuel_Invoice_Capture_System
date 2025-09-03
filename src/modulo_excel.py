"""
Module: modulo_excel.py
Handles Excel operations such as creating shift sheets, saving invoice data,
and calculating totals.

Functions:
    - get_or_create_sheet(filename, fecha_descarga, turno, template_name="Plantilla")
    - save_to_excel(filename, datos, sheet)
    - calculate_totals(filename, sheet)
"""

from openpyxl import Workbook,  load_workbook
import os  

def get_or_create_sheet(filename, fecha_descarga, turno, template_name="Plantilla"):
    """
    Create or retrieve a worksheet for a specific shift based on template.

    Args:
        - filename (str): Path to the Excel file.
        - fecha_descarga (str): Date string in format dd/mm/yyyy.
        - turno (str): Work shift string, "07:00-15:00",
          "15:00-23:00","23:00-07:00".
        - template_name (str): Name of the template sheet to copy (default is "Plantilla").

    Returns:
        - str: Name of the created or existing worksheet.

    Raises:
        - FileNotFoundError: If the Excel file does not exist.
        - ValueError: If the template sheet does not exist.
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"El archivo {filename} no existe. Debe contener una hoja llamada '{template_name}'.")

    wb = load_workbook(filename)
    
    if template_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja plantilla '{template_name}' en el archivo {filename}.")

    # Dynamic sheet name
    sheet = f"{fecha_descarga.replace('/', '-')}_Turno{turno.replace(':', '')}"
    
    # Verify if the sheet already exists
    if sheet not in wb.sheetnames:
        template = wb[template_name]
        ws = wb.copy_worksheet(template)
        ws.title = sheet  

    wb.save(filename)
    return sheet


def save_to_excel(filename, datos, sheet):
    """
    Save extracted invoice data into the specified Excel worksheet.
    
    Args:
        - filename (str): Path to the Excel file.
        - datos (dict): Dictionary containing invoice data.
        - sheet (str): Target worksheet name.

    Returns:
        - int: The row number where the data was inserted.

    Raises:
        - ValueError: If the worksheet does not exist.
    """
    
    wb = load_workbook(filename)
    if sheet not in wb.sheetnames:
        raise ValueError(f"La hoja {sheet} no existe en el archivo {filename}.")

    ws = wb[sheet]

    row = 18
    while ws.cell(row=row, column=2).value not in (None, ""):
        row += 1

    # Mapping data to their corresponding cells
    ws.cell(row=row, column=2, value=datos.get("fecha_carga", ""))       # Column B
    ws.cell(row=row, column=3, value=datos.get("numero_embarque", ""))   # Column C
    ws.cell(row=row, column=4, value=datos.get("clave_equipo", ""))      # Column D
    ws.cell(row=row, column=6, value=datos.get("cantidad_natural", ""))  # Column F
    ws.cell(row=row, column=7, value=datos.get("cantidad_20c", ""))      # Column G

    wb.save(filename)
    return row

def calculate_totals(filename, sheet):
    """
    Calculate the total values for natural liters and liters at 20 °C.

    Args:
        - filename (str): Path to the Excel file.
        - sheet (str): Target worksheet name.

    Returns:
        - tuple: (total_cantidad_natural, total_cantidad_20c) as floats.
    """
    
    wb = load_workbook(filename)
    if sheet not in wb.sheetnames:
        print(f"La hoja {sheet} no existe en el archivo.")
        return

    ws = wb[sheet]
    total_cantidad_natural = 0.0
    total_cantidad_20c = 0.0

    row = 18
    while ws.cell(row=row, column=2).value not in (None, ""):
        try:
            natural = float(ws.cell(row=row, column=6).value)    
            total20 = float(ws.cell(row=row, column=7).value)        
            total_cantidad_natural += natural
            total_cantidad_20c += total20
        except (TypeError, ValueError):
            pass
        row += 1

    return total_cantidad_natural, total_cantidad_20c
